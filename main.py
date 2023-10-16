# Import necessary libraries and modules
from io import BytesIO
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import letter, landscape
import os
import openpyxl
import xlsxwriter
from selenium import webdriver
import tkinter as tk
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import numpy as np
import requests
import urllib.parse
import selenium
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
import time
from selenium.webdriver.support.ui import Select
import warnings

# Ignore DeprecationWarning
warnings.filterwarnings("ignore", category=DeprecationWarning)


# Define a dictionary with form input field names and their corresponding IDs
d = {
    'reg_no': 'ctl00$ContentPlaceHolder1$txtApplicationNumber',
    'name': 'ctl00$ContentPlaceHolder1$txtNameAs12MarkCard',
    'branch': 'ctl00$ContentPlaceHolder1$txtProgramBranch',
    'mobile_num': 'ctl00$ContentPlaceHolder1$txtStudentMobileNumberPresent',
    'email_id': 'ctl00$ContentPlaceHolder1$txtStudentEmailID',
    'dob': 'ctl00$ContentPlaceHolder1$txtDOB',
    'blood_group': 'ctl00$ContentPlaceHolder1$txtBloodGroup',

    # Click on Address Details
    'address_1': 'ctl00$ContentPlaceHolder1$txtPermanentAddressLine1',
    'address_2': 'ctl00$ContentPlaceHolder1$txtPermanentAddressLine2',
    'address_3': 'ctl00$ContentPlaceHolder1$txtPermanentAddressLine3',
    'address_4': 'ctl00$ContentPlaceHolder1$txtPlace',
    'address_5': 'ctl00$ContentPlaceHolder1$txtState',
    'address_6': 'ctl00$ContentPlaceHolder1$txtCountry',
    'address_7': 'ctl00$ContentPlaceHolder1$txtPincode',
    'hostel_block': 'ctl00$ContentPlaceHolder1$txtHostelBlock',
    'hostel_room': 'ctl00$ContentPlaceHolder1$txtHostelRoomNumber',
    # Click on parent details
    'father_name': 'ctl00$ContentPlaceHolder1$txtPFatherName',
    'father_occupation': 'ctl00$ContentPlaceHolder1$txtFatherOccupation',
    'father_email': 'ctl00$ContentPlaceHolder1$txtFatherEmailID',
    'father_mobile': 'ctl00$ContentPlaceHolder1$txtFatherContactNumber',

    'mother_name': 'ctl00$ContentPlaceHolder1$txtPMotherName',
    'mother_occupation': 'ctl00$ContentPlaceHolder1$txtMotherOccupation',
    'mother_email': 'ctl00$ContentPlaceHolder1$txtMotherEmailID',
    'mother_mobile': 'ctl00$ContentPlaceHolder1$txtMotherContactNumber',

    # span: Academics page ID not name
    'roll_no': 'ContentPlaceHolder1_lblRollNo',
    'section': 'ContentPlaceHolder1_lblSection',
}


# Create a dictionary to map Excel cell names to field names
cellMapping = {
    'A6': 'name',
    'D6': 'reg_no',
    'A7': 'section',
    'B7': 'roll_no',
    'D7': 'branch',
    'A8': 'mobile_num',
    'D8': 'email_id',
    'A9': 'hostel_block',
    'B9': 'hostel_room',
    'D9': 'blood_group',
    'E9': 'dob',
    'A12': 'father_name',
    'D12': 'mother_name',
    'A13': 'father_occupation',
    'D13': 'mother_occupation',
    'A17': 'father_mobile',
    'D17': 'mother_mobile',
    'A18': 'father_email',
    'D18': 'mother_email',
    'D20': 'address',
    'A20': 'address'
}

# Define lists of keys for different form pages
page1_keys = ['reg_no', 'name', 'branch',
              'mobile_num', 'email_id', 'dob', 'blood_group']
page2_keys = [
    'address_1',
    'address_2',
    'address_3',
    'address_4',
    'address_5',
    'address_6',
    'address_7',
    'hostel_block',
    'hostel_room',
]
page3_keys = [
    'father_name',
    'father_occupation',
    'father_email',
    'father_mobile',
    'mother_name',
    'mother_occupation',
    'mother_email',
    'mother_mobile'
]

page4_keys = ['roll_no',
              'section']


# Function to get student data from the website
def getStudentData(driver):
    profile_url = 'https://slcm.manipal.edu/StudentProfile.aspx'
    driver.get(profile_url)
    time.sleep(4)
    result = {}
    for key in page1_keys:
        input_element = driver.find_element_by_name(d[key])
        input_value = input_element.get_attribute("value")
        result[key] = input_value
    a_element = driver.find_element_by_xpath(
        "//a[span[text()='Address Details']]")
    a_element.click()
    time.sleep(3)
    for key in page2_keys:
        input_element = driver.find_element_by_name(d[key])
        input_value = input_element.get_attribute("value")
        result[key] = input_value
    a_element = driver.find_element_by_xpath(
        "//a[span[text()='Parent Details']]")
    a_element.click()
    time.sleep(3)
    for key in page3_keys:
        input_element = driver.find_element_by_name(d[key])
        input_value = input_element.get_attribute("value")
        result[key] = input_value
    academics_url = 'https://slcm.manipal.edu/Academics.aspx'
    driver.get(academics_url)
    time.sleep(4)
    a_element = driver.find_element_by_xpath(
        "//a[span[text()='Enrollment Details']]")
    a_element.click()
    time.sleep(3)
    for key in page4_keys:
        span_element = driver.find_element_by_id(d[key])
        span_text = span_element.text
        result[key] = span_text
    result['address'] = '\n' + result['address_1'] + '\n' + result['address_2'] + '\n' + result['address_3'] + '\n' + \
        result['address_4'] + '\n' + result['address_5'] + '\n' + \
        result['address_6'] + '\nPincode: ' + result['address_7']

    del result['address_1']
    del result['address_2']
    del result['address_3']
    del result['address_4']
    del result['address_5']
    del result['address_6']
    del result['address_7']

    return result


# Function to save student data to an Excel file
def saveStudentData(result, regNo):
    workbook = openpyxl.load_workbook('student_details_template.xlsx')
    sheet = workbook['Sheet 1']
    for cell in list(cellMapping.keys()):
        sheet[cell] = sheet[cell].value + ' ' + result[cellMapping[cell]]

    if not os.path.exists(regNo):
        os.makedirs(regNo)

    workbook.save(regNo + '/' + regNo + ' Details.xlsx')


# Function to find and extract data for a student
def findDataForStudent(driver, regNo):
    print("Getting Data for " + regNo)
    # Clicking on 'Student Search'

    a_element = driver.find_element_by_xpath("//a[span[text()='Student Search']]")
    a_element.click()

    # Entering reg_no in input
    input_element = driver.find_element_by_name('ctl00$ContentPlaceHolder1$txtEnrollmentNo')
    input_element.clear()
    input_element.send_keys(regNo)

    element = driver.find_element_by_link_text("Show")
    element.click()

    time.sleep(4)
    result = getStudentData(driver)
    saveStudentData(result, regNo)

    cgpa = ''
    gpaDict = {}
    print("Going to grade sheet page: ")
    gradeUrl = 'https://slcm.manipal.edu/GradeSheet.aspx'
    driver.get(gradeUrl)

    select_element = driver.find_element_by_name(
        "ctl00$ContentPlaceHolder1$ddlSemester")
    options = select_element.find_elements_by_tag_name("option")

    values_to_select = [option.text for option in options]
    values_to_select = values_to_select[:-1]
    values_to_select = [item for item in values_to_select if "&" not in item]
    select_element = Select(select_element)
    time.sleep(2)

    tableId = "ContentPlaceHolder1_grvGradeSheet"

    df = pd.DataFrame(columns=['Sl No.', 'Subject Code', 'Subject Name', 'Actual Semester/Year',
                               'Grade', 'Credit', 'Revaluation1', 'Makeup Exam',
                               'Revaluation2/Revaluation3']
                      )
    for value in values_to_select:
        select_element = Select(driver.find_element_by_name(
            "ctl00$ContentPlaceHolder1$ddlSemester"))
        print("Going to semster: " + value)
        select_element.select_by_value(value)
        time.sleep(4)
        gpaElement = driver.find_element_by_id('ContentPlaceHolder1_lblGPA')
        gpaDict[value] = gpaElement.text
        if(value == 'I'):
            cgpaElement = driver.find_element_by_id(
                'ContentPlaceHolder1_lblCGPA')
            cgpa = cgpaElement.text
        table = driver.find_element_by_xpath(
            "//table[@id='ContentPlaceHolder1_grvGradeSheet']")
        table_html = table.get_attribute('outerHTML')

        df1 = pd.read_html(table_html)[0]
        df = pd.concat([df, df1], axis=0)
        df = df.reset_index(drop=True)

        time.sleep(4)
    df = df.drop(columns=['Sl No.', 'Revaluation1',
                 'Makeup Exam', 'Revaluation2/Revaluation3'])
    print("Going to Back to academics page: ")
    gradeUrl = 'https://slcm.manipal.edu/Academics.aspx'
    driver.get(gradeUrl)
    time.sleep(5)

    print("Clicking on Internal marks sheet")
    a_element = driver.find_element_by_xpath(
        "//a[span[text()='Internal Marks']]")
    a_element.click()

    time.sleep(2)

    print("Getting Select Element")
    select_element = driver.find_element_by_name(
        "ctl00$ContentPlaceHolder1$ddlInternalSemester")
    options = select_element.find_elements_by_tag_name("option")
    values_to_select = [option.text for option in options]
    values_to_select = values_to_select[:-1]
    values_to_select = [item for item in values_to_select if "&" not in item]
    select_element = Select(select_element)
    time.sleep(2)

    text_list = []
    d_list = []
    for value in values_to_select:
        select_element = Select(driver.find_element_by_name(
            "ctl00$ContentPlaceHolder1$ddlInternalSemester"))
        print("Going to semester: " + value)
        select_element.select_by_value(value)
        time.sleep(2)
        element = driver.find_element_by_link_text("Show")
        element.click()
        time.sleep(4)

        h4_elements = driver.find_elements_by_css_selector('h4.panel-text')
        for h4 in h4_elements:
            text = h4.text
            text_list.append(text)

    print("Formating response...")
    for item in text_list:
        code = item[14:22]
        if(len(item.split('Obtained:')) > 1):
            marks = float(item.split('Obtained:')[1][1:6])
            d = {'Subject Code': code, 'Internal Marks': marks}
            d_list.append(d)
    df_internal = pd.DataFrame(d_list)

    result_df = pd.merge(df, df_internal, on='Subject Code', how='outer')
    print("Got data for " + regNo)
    return [result_df, gpaDict, cgpa]


# Function to create a PDF report from the Excel data
def createPdf(fileName):
    df = pd.read_excel(fileName+'.xlsx')
    df = df.fillna('')
    df = df.drop(df.columns[:2], axis=1)
    pdf_file = fileName+'.pdf'

    pdf = SimpleDocTemplate(pdf_file, pagesize=landscape(letter))

    elements = []

    text1 = "Office of Quality Assurance"
    text2 = "Department of Computer Science"
    text3 = 'Semester ' + \
        fileName.split('Semester')[1].split('result')[0] + 'Result'
    styles = getSampleStyleSheet()
    text1_style = styles['Normal']
    text1_style.fontSize = 12  #
    text1_paragraph = Paragraph(text1, text1_style)
    elements.append(text1_paragraph)

    elements.append(Spacer(1, 0.25 * inch))

    text2_style = styles['Title']
    text2_style.fontSize = 14
    text2_style.alignment = 1
    text2_paragraph = Paragraph(text2, text2_style)
    elements.append(text2_paragraph)

    text3_style = styles['Title']
    text3_style.fontSize = 12
    text3_style.alignment = 1
    text3_paragraph = Paragraph(text3, text3_style)
    elements.append(text3_paragraph)

    data = df.values.tolist()[2:]
    for rowInd in range(len(data)):
        row = data[rowInd]
        if(rowInd != 0):
            last_column_width = row[-1]
            row[-1] = 4 * last_column_width
        else:
            last_column_width = row[-1]
            row[-1] = '    '*2 + last_column_width + '    '*2
    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),  # Table background color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center-align all cells
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Header padding
        # Table body background color
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Table grid
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Body font
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Left-align all cells
        ('LEFTPADDING', (0, 0), (-1, -1), 6),  # Cell padding
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),  # Cell padding
        ('COLWIDTH', (-1, -1), (-1, -1), 10 * inch),
    ])

    table.setStyle(style)

    elements.append(table)

    elements.append(Spacer(1, 0.25 * inch))
    red_title_style = ParagraphStyle(
        name='RedTitleStyle', parent=styles['Title'])
    red_title_style.textColor = colors.red

    title = Paragraph(
        "<font size='12' color='red'>Record of Mentor Mentee Meeting</font>", red_title_style)
    elements.append(title)

    second_table_data = [["Date and Time", "Discussion",
                          "Signature of Student"], ["", "", ""]]

    second_table = Table(second_table_data, colWidths=[3*inch, 5*inch, 2*inch])
    second_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('COLWIDTH', (-1, -1), (-1, -1), 10 * inch),
    ]))

    padding_y = 0.5 * inch
    for i in range(len(second_table_data[1])):
        second_table.setStyle(TableStyle([
            ('BOTTOMPADDING', (i, 1), (i, 1), padding_y)
        ]))

    elements.append(Spacer(1, 0.25 * inch))
    elements.append(second_table)

    pdf.build(elements)

    print(f'PDF saved to {pdf_file}')


# Function to create Excel and PDF reports for a student
def create(regNo, df, sem, gpaDict, cgpa):
    df_columns = ['Subject Code', 'Subject Name',
                  'Credit', 'Internal Marks', 'Grade']
    workbook = openpyxl.load_workbook('template.xlsx')
    sheet = workbook['Sheet1']
    df_filtered = df[df['Actual Semester/Year'] == sem]
    sheet['C3'] = 'Semester ' + sem + ' End Semester Results'
    sheet['D15'] = 'GPA: ' + gpaDict[sem]
    sheet['F15'] = 'CGPA: ' + cgpa
    for i in range(len(df_columns)):
        start_cell = sheet.cell(row=5, column=i+3)
        for index, code in enumerate(df_filtered[df_columns[i]], start=1):
            current_cell = start_cell.offset(row=index - 1, column=0)
            current_cell.value = code
    if not os.path.exists(regNo):
        os.makedirs(regNo)
    fileName = regNo + '/' + regNo + ' Semester '+sem+' result'
    workbook.save(fileName+'.xlsx')
    createPdf(fileName)
    if os.path.exists(fileName + '.xlsx'):
        os.remove(fileName + '.xlsx')

    print(fileName + " saved successfully")


# Main function to scrape the website
def scrape_website():
    input_text = entry.get("1.0", "end-1c")
    PATH = 'chromedriver.exe'
    students = input_text.split('\n')
    print(students)
    chrome_service = ChromeService(executable_path=PATH)
    url = 'https://slcm.manipal.edu/'
    driver = webdriver.Chrome(service=chrome_service)
    driver.get(url)

    print("Waiting for login..")
    time.sleep(20)
    time.sleep(20)
    for reg_no in students:
        df, gpaDict, cgpa = findDataForStudent(driver, reg_no)
        df = df.dropna()
        l = list(set(df['Actual Semester/Year']))
        l.sort()
        print("Saving data for " + reg_no)
        for sem in l:
            create(reg_no, df, sem, gpaDict, cgpa)
        print("Going back to search page: ")
        driver.get("https://slcm.manipal.edu/FacultyHome.aspx")
        time.sleep(2)
    # To go back to faculty home page and repeat process for all students
    driver.quit()


# Create a GUI window
window = tk.Tk()
window.title("Mentor Form Web Scrapping")

window.geometry("400x150")
window.configure(bg='#F0F0F0')

label = tk.Label(window, text="Enter registration numbers, 1 on each line:", font=(
    "Arial", 11), bg='#F0F0F0')
label.pack(pady=10)

entry = tk.Text(window,  height=5, font=("Arial", 12))
entry.pack(pady=5)

submit_button = tk.Button(window, text="Submit",
                          command=scrape_website, font=("Arial", 12))
submit_button.pack(pady=10)
window.geometry("400x200")

window.mainloop()